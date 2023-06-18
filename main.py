import subprocess
import xml.etree.ElementTree as ET
import tkinter as tk
from tkinter import *
from tkinter import filedialog, messagebox, ttk
import getpass
import json
import xml.dom.minidom as minidom
from openpyxl import Workbook, load_workbook


def extract_cpu_info(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    cpu_info = {}
    for subnode in root.iter('SubNode'):
        for property_node in subnode.iter('Property'):
            entry = property_node.find('Entry').text
            description = property_node.find('Description').text
            if entry == 'CPU Brand Name':
                cpu_info['CPU Brand Name'] = description

    return cpu_info


def extract_monitor_info(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    monitor_info = []
    for subnode in root.iter('SubNode'):
        for property_node in subnode.iter('Property'):
            entry = property_node.find('Entry').text
            description = property_node.find('Description').text
            if entry == 'Monitor Name':
                monitor_info.append(description)

    return monitor_info


def extract_gpu_info(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    gpu_info = {}
    for subnode in root.iter('SubNode'):
        for property_node in subnode.iter('Property'):
            entry = property_node.find('Entry').text
            description = property_node.find('Description').text
            if entry == 'Video Chipset':
                gpu_info['Video Chipset'] = description

    return gpu_info


def extract_ram_info(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    ram_info = {}
    for memory_node in root.iter('MEMORY'):
        node_name = memory_node.find('NodeName').text.strip()
        if node_name == 'Memory':
            for property_node in memory_node.iter('Property'):
                entry = property_node.find('Entry').text
                description = property_node.find('Description').text
                if entry == 'Total Memory Size':
                    ram_info['Total Memory Size'] = description

    return ram_info


def extract_motherboard_info(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    motherboard_info = {}
    for motherboard_node in root.iter('MOBO'):
        node_name = motherboard_node.find('NodeName').text.strip()
        if node_name == 'Motherboard':
            for property_node in motherboard_node.iter('Property'):
                entry = property_node.find('Entry').text
                description = property_node.find('Description').text
                if entry == 'Motherboard Model':
                    motherboard_info['Motherboard Model'] = description

    return motherboard_info


def extract_memory_size_info(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    memory_size_info = []
    for subnode in root.iter('SubNode'):
        for property_node in subnode.iter('Property'):
            entry = property_node.find('Entry').text
            description = property_node.find('Description').text
            if entry == 'Drive Capacity' and description not in memory_size_info:
                memory_size_info.append(description)

    return memory_size_info




def extract_username_info(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    username_info = {}
    for memory_node in root.iter('COMPUTER'):
        for property_node in memory_node.iter('Property'):
            entry = property_node.find('Entry').text
            description = property_node.find('Description').text
            if entry == 'Current User Name':
                username_info['Current User Name'] = description

    return username_info


def extract_memory_type_info(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    memory_type_info = []
    for subnode in root.iter('SubNode'):
        for property_node in subnode.iter('Property'):
            entry = property_node.find('Entry').text
            description = property_node.find('Description').text
            if entry == 'Media Rotation Rate' and description not in memory_type_info:
                memory_type_info.append(description)

    return memory_type_info


def get_domain_info():
    try:
        output = subprocess.check_output(['wmic', 'computersystem', 'get', 'domain'], universal_newlines=True)
        lines = output.splitlines()
        for line in lines:
            if line.strip() and not line.startswith('Domain'):
                domain_info = line.strip()
                return domain_info
    except subprocess.CalledProcessError:
        pass

    return None


def load_template(window, fields):
    try:
        with open('template.json', 'r') as file:
            template = json.load(file)
    except FileNotFoundError:
        return

    for field_name in template:
        field_frame = tk.Frame(window)
        field_frame.pack(anchor='w')

        field_label = tk.Label(field_frame, text=field_name + ':')
        field_label.pack(side='top')

        field_var = tk.StringVar()
        field_entry = tk.Entry(field_frame, textvariable=field_var, state='normal', width=30)
        field_entry.pack(side='left')

        delete_button = tk.Button(field_frame, text='-', command=lambda: delete_custom_field(fields, field_frame))
        delete_button.configure(bg='#a60000', cursor='hand2', fg='#f0f0f0', font=('Arial', 12, 'bold'), relief='flat')
        delete_button.pack(side='left')

        fields.append(field_frame)


def add_custom_field(window, fields):
    field_name = tk.simpledialog.askstring('Custom Field', 'Enter field name:')
    if field_name:
        field_frame = tk.Frame(window)
        field_frame.pack(anchor='w')

        field_label = tk.Label(field_frame, text=field_name + ':')
        field_label.pack(side='top')

        field_var = tk.StringVar()
        field_entry = tk.Entry(field_frame, textvariable=field_var, state='normal', width=30)
        field_entry.pack(side='left')

        delete_button = tk.Button(field_frame, text='-', command=lambda: delete_custom_field(fields, field_frame))
        delete_button.configure(bg='#a60000', cursor='hand2', fg='#f0f0f0', font=('Arial', 12, 'bold'), relief='flat')
        delete_button.pack(side='left')

        fields.append(field_frame)

        save_template(fields)


def save_template(fields):
    template = []
    for field_frame in fields:
        field_name = field_frame.winfo_children()[0]['text'][:-1]
        template.append(field_name)

    with open('template.json', 'w') as file:
        json.dump(template, file)


def delete_custom_field(fields, field_frame):
    result = messagebox.askquestion('Delete Field', 'Are you sure you want to delete this field?', icon='warning')
    if result == 'yes':
        field_label = field_frame.winfo_children()[0]
        field_name = field_label['text'][:-1]
        field_frame.destroy()
        fields.remove(field_frame)
        update_template_file(fields, field_name)


def update_template_file(fields, deleted_field_name):
    template = []
    for field_frame in fields:
        field_label = field_frame.winfo_children()[0]
        field_name = field_label['text'][:-1]
        template.append(field_name)

    with open('template.json', 'r') as file:
        template_data = json.load(file)

    template_data.remove(deleted_field_name)

    with open('template.json', 'w') as file:
        json.dump(template_data, file)


def select_xml_file():
    file_path = filedialog.askopenfilename(filetypes=[('XML files', '*.xml')])
    if file_path:
        cpu_info = extract_cpu_info(file_path)
        gpu_info = extract_gpu_info(file_path)
        monitor_info = extract_monitor_info(file_path)
        ram_info = extract_ram_info(file_path)
        motherboard_info = extract_motherboard_info(file_path)
        memory_size_info = extract_memory_size_info(file_path)
        memory_type_info = extract_memory_type_info(file_path)
        domain_info = get_domain_info()
        username_info = extract_username_info(file_path)

        data = {
            'Processor': cpu_info.get('CPU Brand Name'),
            'Graphics card': gpu_info.get('Video Chipset'),
            'Monitors': ': '.join(monitor_info),
            'RAM': ram_info.get('Total Memory Size'),
            'Motherboard': motherboard_info.get('Motherboard Model'),
            'Domain': domain_info,
            'Memory capacity': ': '.join(memory_size_info),
            'Memory type': ': '.join(memory_type_info),
            'Username': username_info.get('Current User Name')
        }

        create_info_window(cpu_info, gpu_info, monitor_info, ram_info, motherboard_info, domain_info,
                           memory_size_info, memory_type_info, username_info, data, file_path)


def save(xml_file, data, custom_fields):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    # Create a new node for the saved information
    info_node = ET.Element('SavedInfo')
    for key, value in data.items():
        property_node = ET.SubElement(info_node, 'Property')
        entry_node = ET.SubElement(property_node, 'Entry')
        entry_node.text = key
        description_node = ET.SubElement(property_node, 'Description')
        description_node.text = value

    # Add data from custom fields
    for field_frame in custom_fields:
        field_entry = field_frame.winfo_children()[1]  # Access the Entry widget within the frame
        field_label = field_frame.winfo_children()[0]['text'][:-1]
        if field_entry.get() != '':
            property_node = ET.SubElement(info_node, 'Property')
            entry_node = ET.SubElement(property_node, 'Entry')
            entry_node.text = field_label
            description_node = ET.SubElement(property_node, 'Description')
            description_node.text = field_entry.get()

    # Append the new node to the root
    root.append(info_node)

    # Save the modified XML file
    xml_string = ET.tostring(root, encoding='utf-8')
    dom = minidom.parseString(xml_string)
    pretty_xml_string = dom.toprettyxml(indent='  ')

    with open(xml_file, 'w', encoding='utf-8') as file:
        file.write(pretty_xml_string)

    # Save the data to Excel file
    excel_file = 'All_extracted_data.xlsx'
    if not data_exists(excel_file):
        create_excel_file(excel_file)

    save_data_to_excel(excel_file, data, custom_fields)


def data_exists(excel_file):
    try:
        workbook = load_workbook(excel_file)
        return True
    except FileNotFoundError:
        return False


def create_excel_file(excel_file):
    workbook = Workbook()
    sheet = workbook.active

    headers = [
        'Processor', 'Graphics card', 'Monitors', 'RAM', 'Motherboard', 'Domain',
        'Memory capacity', 'Memory type', 'Username'
    ]

    sheet.append(headers)
    workbook.save(excel_file)


# Function to load the current_row value from a JSON file
def load_current_row():
    try:
        with open("current_row.json", "r") as file:
            data = json.load(file)
            return data["current_row"]
    except FileNotFoundError:
        return 0


# Function to save the current_row value to a JSON file
def save_current_row(current_row):
    data = {"current_row": current_row}
    with open("current_row.json", "w") as file:
        json.dump(data, file)


def save_data_to_excel(excel_file, data, custom_fields):
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    # Find the next available row
    last_row = sheet.max_row
    current_row = last_row + 1
    save_current_row(current_row)

    row_data = [
        data.get('Processor'),
        data.get('Graphics card'),
        data.get('Monitors'),
        data.get('RAM'),
        data.get('Motherboard'),
        data.get('Domain'),
        data.get('Memory capacity'),
        data.get('Memory type'),
        data.get('Username')
    ]

    # Get the last column index with data
    last_column = sheet.max_column

    # Add the custom field labels to the header row if they don't exist
    for i, field_frame in enumerate(custom_fields):
        field_label = field_frame.winfo_children()[0]['text'][:-1]
        header_exists = False
        for column in range(1, last_column + 1):
            if sheet.cell(row=1, column=column).value == field_label:
                header_exists = True
                break
        if not header_exists:
            sheet.cell(row=1, column=last_column + i + 1).value = field_label

    # Update the last column index with data
    last_column = sheet.max_column

    # Add the data from custom fields
    for i, field_frame in enumerate(custom_fields):
        field_label = field_frame.winfo_children()[0]['text'][:-1]
        field_entry = field_frame.winfo_children()[1]  # Access the Entry widget within the frame
        description = field_entry.get()
        column_index = None
        for column in range(1, last_column + 1):
            if sheet.cell(row=1, column=column).value == field_label:
                column_index = column
                break
        if column_index is None:
            column_index = last_column + i + 1
            sheet.cell(row=1, column=column_index).value = field_label
        sheet.cell(row=current_row, column=column_index).value = description

    # Add the row data
    for i, value in enumerate(row_data):
        sheet.cell(row=current_row, column=i + 1).value = value

    # Save the modified Excel file
    workbook.save(excel_file)


def create_info_window(cpu_info, gpu_info, monitor_info, ram_info, motherboard_info, domain_info, memory_size_info,
                       memory_type_info, username_info, data, xml_file):
    window = tk.Toplevel()
    window.title('System Information')
    window.geometry('200x900')
    window.resizable(False, True)

    # Set the alignment to the left
    window.grid_propagate(False)
    # Custom fields
    custom_fields = []
    custom_fields_frame = tk.Frame(window)
    custom_fields_frame.pack(anchor='center')

    add_button = tk.Button(custom_fields_frame, text='+', command=lambda: add_custom_field(window, custom_fields))
    add_button.configure(bg='#008a00', cursor='hand2', fg='#f0f0f0', font=('Arial', 12, 'bold'), relief='flat')
    add_button.pack(side='left', padx='10')

    save_button = tk.Button(custom_fields_frame, text='Save', command=lambda: save(xml_file, data, custom_fields))
    save_button.configure(bg='#008a00', cursor='hand2', fg='#f0f0f0', font=('Arial', 12, 'bold'), relief='flat')
    save_button.pack(side='left', padx='10')

    for field_frame in custom_fields:
        field_frame.pack(anchor='w')

    # CPU
    tk.Label(window, text='Processor:').pack(anchor='w')
    processor_name_var = tk.StringVar()
    processor_name_var.set(cpu_info.get('CPU Brand Name'))
    processor_name_entry = tk.Entry(window, textvariable=processor_name_var, state='normal', width=30)
    processor_name_entry.pack(anchor='w')

    # GPU
    tk.Label(window, text='Graphics card:').pack(anchor='w')
    video_controller_var = tk.StringVar()
    video_controller_var.set(gpu_info.get('Video Chipset'))
    video_controller_entry = tk.Entry(window, textvariable=video_controller_var, state='normal', width=30)
    video_controller_entry.pack(anchor='w')

    # Monitor
    for i, monitor_name in enumerate(monitor_info):
        tk.Label(window, text=f'Monitor {i + 1}:').pack(anchor='w')
        monitor_var = tk.StringVar()
        monitor_var.set(monitor_name)
        monitor_entry = tk.Entry(window, textvariable=monitor_var, state='normal', width=30)
        monitor_entry.pack(anchor='w')

    # RAM memory
    tk.Label(window, text='RAM:').pack(anchor='w')
    ram_memory_var = tk.StringVar()
    ram_memory_var.set(ram_info.get('Total Memory Size'))
    ram_memory_entry = tk.Entry(window, textvariable=ram_memory_var, state='normal', width=30)
    ram_memory_entry.pack(anchor='w')

    # Motherboard
    tk.Label(window, text='Motherboard:').pack(anchor='w')
    motherboard_var = tk.StringVar()
    motherboard_var.set(motherboard_info.get('Motherboard Model'))
    motherboard_entry = tk.Entry(window, textvariable=motherboard_var, state='normal', width=30)
    motherboard_entry.pack(anchor='w')

    # Domain
    tk.Label(window, text='Domain:').pack(anchor='w')
    domain_var = tk.StringVar()
    domain_var.set(domain_info)
    domain_entry = tk.Entry(window, textvariable=domain_var, state='normal', width=30)
    domain_entry.pack(anchor='w')

    # Memory size
    for i, memory_size in enumerate(memory_size_info):
        tk.Label(window, text=f'Drive size {i + 1}:').pack(anchor='w')
        memory_size_var = tk.StringVar()
        memory_size_var.set(memory_size)
        memory_size_entry = tk.Entry(window, textvariable=memory_size_var, state='normal', width=30)
        memory_size_entry.pack(anchor='w')

    # Memory type
    for i, memory_type in enumerate(memory_type_info):
        tk.Label(window, text=f'Memory type {i + 1}:').pack(anchor='w')
        memory_type_var = tk.StringVar()
        memory_type_var.set(memory_type)
        memory_type_entry = tk.Entry(window, textvariable=memory_type_var, state='normal', width=30)
        memory_type_entry.pack(anchor='w')

    # Username
    tk.Label(window, text='Username:').pack(anchor='w')
    username_var = tk.StringVar()
    username_var.set(username_info.get('Current User Name'))
    username_entry = tk.Entry(window, textvariable=username_var, state='normal', width=30)
    username_entry.pack(anchor='w')

    load_template(window, custom_fields)


root = tk.Tk()
root.withdraw()
current_row = load_current_row()
select_xml_file()
root.mainloop()
